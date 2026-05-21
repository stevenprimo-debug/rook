#!/usr/bin/env python3
"""
Generic Shopify operate-mode helper. Sanitized from a prior client engagement.

Pulls orders from Shopify Admin API, exports to Excel.
Captures line_items[].properties[] that CSV export drops.

Configuration via environment variables (see .env.example):
  SHOPIFY_STORE_HANDLE      — myshopify.com subdomain handle
  SHOPIFY_CLIENT_ID         — OAuth app client ID
  SHOPIFY_CLIENT_SECRET     — OAuth app client secret
  SHOPIFY_CUSTOM_PROPERTY   — Line-item property name for merchant-specific ID field
                              (e.g. "House number", "Unit", "Reference")
"""

import os
import sys
import time
import argparse
from datetime import datetime
from typing import Optional, Dict, List, Any
from collections import defaultdict
from urllib.parse import urlencode

import requests
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv


load_dotenv()

# Constants — all config-driven, no hardcoded merchant data
STORE_HANDLE = os.environ.get("SHOPIFY_STORE_HANDLE", "")
API_VERSION = os.environ.get("SHOPIFY_API_VERSION", "2024-01")
API_BASE = f"https://{STORE_HANDLE}.myshopify.com/admin/api/{API_VERSION}"
ORDERS_ENDPOINT = f"{API_BASE}/orders.json"
RATE_LIMIT_THRESHOLD = 35  # Sleep if X-Shopify-Shop-Api-Call-Limit >= 35/40

# Merchant-specific line-item property name — read from config, not hardcoded
CUSTOM_PROPERTY = os.environ.get("SHOPIFY_CUSTOM_PROPERTY", "Custom field")
# Properties to exclude from the "extra properties" columns (noise)
IGNORE_PROPERTIES_RAW = os.environ.get("SHOPIFY_IGNORE_PROPERTIES", "Lead time acknowledgment")
IGNORE_PROPERTIES = {p.strip() for p in IGNORE_PROPERTIES_RAW.split(",")}


def get_auth_token() -> str:
    """
    Get a fresh Admin API token via client_credentials grant.
    Token expires in 24h; we fetch a new one each run to stay current.
    Requires SHOPIFY_CLIENT_ID and SHOPIFY_CLIENT_SECRET in .env.
    """
    client_id = os.getenv("SHOPIFY_CLIENT_ID")
    client_secret = os.getenv("SHOPIFY_CLIENT_SECRET")
    store = os.getenv("SHOPIFY_STORE_HANDLE", STORE_HANDLE)

    if not client_id or not client_secret:
        raise ValueError(
            "SHOPIFY_CLIENT_ID and SHOPIFY_CLIENT_SECRET must be set in .env"
        )

    url = f"https://{store}.myshopify.com/admin/oauth/access_token"
    resp = requests.post(url, json={
        "client_id": client_id,
        "client_secret": client_secret,
        "grant_type": "client_credentials",
    })
    resp.raise_for_status()
    data = resp.json()
    token = data.get("access_token")
    if not token:
        raise ValueError(f"No access_token in response: {data}")
    print(f"  Auth: token acquired (scope: {data.get('scope', 'unknown')})")
    return token


def get_headers(token: str) -> Dict[str, str]:
    """Build request headers."""
    return {
        "X-Shopify-Access-Token": token,
        "Content-Type": "application/json",
    }


def check_rate_limit(response: requests.Response) -> None:
    """Check X-Shopify-Shop-Api-Call-Limit header; sleep if necessary."""
    limit_header = response.headers.get("X-Shopify-Shop-Api-Call-Limit")
    if limit_header:
        current, total = map(int, limit_header.split("/"))
        if current >= RATE_LIMIT_THRESHOLD:
            sleep_time = 5
            print(f"  [RATE] {current}/{total} calls used. Sleeping {sleep_time}s...")
            time.sleep(sleep_time)


def fetch_orders(
    token: str,
    dry_run: bool = False,
    since_id: Optional[int] = None,
    created_at_min: Optional[str] = None,
    created_at_max: Optional[str] = None,
) -> List[Dict[str, Any]]:
    """
    Fetch all orders from Shopify Admin API using cursor pagination.

    Args:
        token: Shopify Admin API access token
        dry_run: If True, fetch only first page (250 orders)
        since_id: Return only orders with ID > since_id (incremental pull)
        created_at_min: Return only orders created >= this date (ISO 8601)
        created_at_max: Return only orders created <= this date (ISO 8601)

    Returns:
        List of order dictionaries

    NOTE: cursor pages use ONLY page_info + limit — do not add filter params
    to cursor-page requests (causes HTTP 422). See KNOWN_ISSUES.md KI-003.
    """
    headers = get_headers(token)
    orders = []
    page_info = None  # Shopify cursor pagination token
    page_count = 0

    while True:
        page_count += 1

        if page_info:
            # Cursor page: ONLY page_info + limit allowed (no other filters)
            url = f"{ORDERS_ENDPOINT}?page_info={page_info}&limit=250"
        else:
            # First page: full filter params
            params = {
                "status": "any",
                "limit": 250,
                "fields": (
                    "id,order_number,name,created_at,updated_at,"
                    "financial_status,fulfillment_status,"
                    "customer,shipping_address,"
                    "line_items,shipping_lines,fulfillments"
                ),
            }
            if since_id:
                params["since_id"] = since_id
            if created_at_min:
                params["created_at_min"] = created_at_min
            if created_at_max:
                params["created_at_max"] = created_at_max
            url = f"{ORDERS_ENDPOINT}?{urlencode(params)}"

        print(f"[PAGE {page_count}] {url[:120]}")

        response = requests.get(url, headers=headers)
        response.raise_for_status()

        check_rate_limit(response)

        data = response.json()
        page_orders = data.get("orders", [])
        orders.extend(page_orders)

        print(f"  Fetched {len(page_orders)} orders. Total so far: {len(orders)}")

        if dry_run:
            print("[DRY-RUN] Stopping after first page.")
            break

        # Extract page_info cursor from Link header
        # Shopify format: <https://store.myshopify.com/...?page_info=XXX&limit=250>; rel="next"
        link_header = response.headers.get("Link", "")
        page_info = None
        if link_header:
            for link in link_header.split(","):
                if 'rel="next"' in link:
                    url_start = link.find("<") + 1
                    url_end = link.find(">")
                    if url_start > 0 and url_end > url_start:
                        next_url = link[url_start:url_end]
                        if "page_info=" in next_url:
                            pi_start = next_url.find("page_info=") + 10
                            pi_end = next_url.find("&", pi_start)
                            page_info = next_url[pi_start:] if pi_end == -1 else next_url[pi_start:pi_end]
                    break

        if not page_info:
            break

    return orders


def extract_custom_field(order: Dict[str, Any]) -> tuple[str, str, str]:
    """
    Extract the merchant-configured custom line-item property (e.g. house number,
    unit number, reference code).

    Returns (property_value, shipping_address_token, mismatch_flag).

    - property_value: value of CUSTOM_PROPERTY from line-item properties (authoritative)
    - shipping_token: first token of shipping_address.address1 (fallback cross-check)
    - mismatch_flag: 'MISMATCH' if both present and different, else ''
    """
    prop_value = ""
    for item in order.get("line_items", []):
        for prop in item.get("properties", []):
            if prop.get("name", "").strip().lower() == CUSTOM_PROPERTY.lower():
                prop_value = str(prop.get("value", "")).strip()
                break
        if prop_value:
            break

    shipping = order.get("shipping_address", {}) or {}
    address1 = shipping.get("address1", "") or ""
    # First token of address line (e.g. street number)
    ship_value = address1.split()[0] if address1.strip() else ""

    mismatch = ""
    if prop_value and ship_value and prop_value != ship_value:
        mismatch = "MISMATCH"

    return prop_value, ship_value, mismatch


def extract_properties(order: Dict[str, Any]) -> Dict[str, str]:
    """
    Extract all unique property keys and values from an order's line items.
    Skips ignored properties (noise like lead-time acknowledgment).
    Returns a dict mapping property_name -> property_value (concatenated if multiple).
    """
    properties_dict = defaultdict(list)

    for item in order.get("line_items", []):
        for prop in item.get("properties", []):
            name = prop.get("name", "")
            value = prop.get("value", "")
            if name and value and name not in IGNORE_PROPERTIES:
                properties_dict[name].append(str(value))

    # Flatten: join multiple values with "; "
    return {k: "; ".join(v) for k, v in properties_dict.items()}


def build_inventory(orders: List[Dict[str, Any]]) -> Dict[str, int]:
    """
    Build inventory of all unique property keys and their frequency.
    """
    inventory = defaultdict(int)

    for order in orders:
        props = extract_properties(order)
        for key in props.keys():
            inventory[key] += 1

    return dict(inventory)


def _order_to_row(order: Dict[str, Any]) -> List[Any]:
    """
    Convert one order dict to an Excel-ready row.
    One row per ORDER (not per line item).

    Columns mirror the Shopify orders admin UI.
    """
    customer = order.get("customer", {}) or {}
    line_items = order.get("line_items", [])
    shipping_lines = order.get("shipping_lines", []) or []
    fulfillments = order.get("fulfillments", []) or []

    # Order #
    order_num = f"#{order.get('order_number', '')}"

    # Date
    created = order.get("created_at", "")
    date_str = ""
    if created:
        # Shopify format: 2026-03-20T14:32:00-05:00 → MM/DD/YYYY
        date_str = f"{created[5:7]}/{created[8:10]}/{created[:4]}"

    # Customer name
    first = customer.get("first_name", "") or ""
    last = customer.get("last_name", "") or ""
    customer_name = f"{first} {last}".strip()

    # Payment Status — flag chargebacks
    financial = order.get("financial_status", "") or ""
    payment_status = "Chargeback" if "chargeback" in financial.lower() else ""

    # Fulfillment Status
    fulfillment = order.get("fulfillment_status") or "unfulfilled"

    # Items — total quantity across all line items
    item_count = sum(item.get("quantity", 1) for item in line_items)

    # Delivery Status — from most recent fulfillment's shipment_status
    delivery_status = ""
    if fulfillments:
        latest = fulfillments[-1]
        shipment_status = latest.get("shipment_status") or ""
        delivery_status = shipment_status.replace("_", " ").title() if shipment_status else ""

    # Delivery Method — from first shipping line
    delivery_method = ""
    if shipping_lines:
        delivery_method = shipping_lines[0].get("title", "") or ""

    # Custom field — from line-item property (authoritative)
    custom_val, _, _ = extract_custom_field(order)

    return [
        order_num,
        date_str,
        customer_name,
        payment_status,
        fulfillment,
        item_count,
        delivery_status,
        delivery_method,
        custom_val,
    ]


SHOPIFY_HEADERS = [
    "Order #",
    "Date",
    "Customer",
    "Payment Status",
    "Fulfillment Status",
    "Items",
    "Delivery Status",
    "Delivery Method",
    CUSTOM_PROPERTY,
]

# Fixed column widths
SHOPIFY_COL_WIDTHS = [10, 12, 24, 16, 18, 7, 16, 26, 14]


def _write_sheet(
    ws,
    orders: List[Dict[str, Any]],
    header_font: Font,
) -> None:
    """Write SHOPIFY_HEADERS + data rows to a worksheet."""
    ws.append(SHOPIFY_HEADERS)

    hdr_fill = PatternFill(start_color="1C1C1C", end_color="1C1C1C", fill_type="solid")
    hdr_font_white = Font(bold=True, color="FFFFFF", name=header_font.name)
    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font_white
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

    # Zebra stripe row fills
    stripe_a = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    stripe_b = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    missing_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for i, order in enumerate(orders):
        row_data = _order_to_row(order)
        ws.append(row_data)

        excel_row = ws[ws.max_row]
        base_fill = stripe_a if i % 2 == 0 else stripe_b
        custom_val = row_data[8]  # index 8 = custom field

        for j, cell in enumerate(excel_row):
            # Highlight missing custom field value
            if j == 8 and not custom_val:
                cell.fill = missing_fill
            else:
                cell.fill = base_fill
            cell.alignment = Alignment(vertical="center", wrap_text=False)

    # Column widths
    for col_idx, width in enumerate(SHOPIFY_COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 20


def export_to_excel(
    orders: List[Dict[str, Any]],
    property_inventory: Dict[str, int],
    output_path: str,
) -> None:
    """
    Export orders to Excel — two sheets:
      Sheet 1 "Unfulfilled"  — orders where fulfillment_status is null / "unfulfilled"
      Sheet 2 "All Orders"   — every order
    Columns mirror the Shopify orders admin UI.
    """
    wb = Workbook()
    header_font = Font(bold=True)

    # Split into unfulfilled vs. all
    unfulfilled = [o for o in orders if not o.get("fulfillment_status")]

    # Sheet 1: Unfulfilled (active default view)
    ws_unfulfilled = wb.active
    ws_unfulfilled.title = "Unfulfilled"
    _write_sheet(ws_unfulfilled, unfulfilled, header_font)

    # Sheet 2: All orders
    ws_all = wb.create_sheet("All Orders")
    _write_sheet(ws_all, orders, header_font)

    wb.save(output_path)
    print(f"\nExported to: {output_path}")
    print(f"  Unfulfilled sheet: {len(unfulfilled)} orders")
    print(f"  All Orders sheet:  {len(orders)} orders")


def main():
    parser = argparse.ArgumentParser(
        description="Pull Shopify orders and export to Excel."
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Fetch only first page (250 orders) for testing.",
    )
    parser.add_argument(
        "--since-id",
        type=int,
        help="Return only orders with ID > since_id (incremental pull).",
    )
    parser.add_argument(
        "--created-at-min",
        type=str,
        help="Return only orders created >= this date (ISO 8601, e.g., 2024-01-01).",
    )
    parser.add_argument(
        "--created-at-max",
        type=str,
        help="Return only orders created <= this date (ISO 8601, e.g., 2024-02-18).",
    )
    parser.add_argument(
        "--output",
        type=str,
        default=None,
        help="Output Excel file path (default: orders_<timestamp>.xlsx in current dir).",
    )

    args = parser.parse_args()

    if not STORE_HANDLE:
        print("ERROR: SHOPIFY_STORE_HANDLE env var not set.", file=sys.stderr)
        sys.exit(1)

    # Get token
    try:
        token = get_auth_token()
    except ValueError as e:
        print(f"ERROR: {e}", file=sys.stderr)
        sys.exit(1)

    # Fetch orders
    print(f"[START] Fetching orders from {STORE_HANDLE}...\n")
    try:
        orders = fetch_orders(
            token,
            dry_run=args.dry_run,
            since_id=args.since_id,
            created_at_min=args.created_at_min,
            created_at_max=args.created_at_max,
        )
    except requests.exceptions.RequestException as e:
        print(f"ERROR: API request failed: {e}", file=sys.stderr)
        sys.exit(1)

    print(f"\n[DONE] Fetched {len(orders)} total orders.\n")

    # Build property inventory
    property_inventory = build_inventory(orders)

    print("[PROPERTIES] Unique property keys found:")
    if property_inventory:
        for key in sorted(property_inventory.keys(), key=lambda k: -property_inventory[k]):
            print(f"  {key}: {property_inventory[key]} orders")

        orders_with_props = sum(
            1 for order in orders if extract_properties(order)
        )
        print(f"\nOrders with at least one property: {orders_with_props}/{len(orders)}\n")
    else:
        print("  (none found)\n")

    # Export to Excel
    if args.output is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        args.output = f"orders_{timestamp}.xlsx"

    export_to_excel(orders, property_inventory, args.output)

    unfulfilled_count = sum(1 for o in orders if not o.get("fulfillment_status"))
    print(f"\n[SUCCESS] Export complete.")
    print(f"  Total orders:  {len(orders)}")
    print(f"  Unfulfilled:   {unfulfilled_count}")
    print(f"  Output: {args.output}")


if __name__ == "__main__":
    main()
