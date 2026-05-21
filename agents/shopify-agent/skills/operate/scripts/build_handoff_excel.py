#!/usr/bin/env python3
"""
Generic Shopify operate-mode helper. Sanitized from a prior client engagement.

Build a merchant handoff Excel — master merged with production-status coloring.

Color legend is fully config-driven via agents/shopify-agent/config/status_color_map.json.
No merchant-specific color logic is hardcoded here.

Configuration via environment variables (see .env.example):
  SHOPIFY_VAULT_ROOT       — base path for this merchant's data folder
  SHOPIFY_COLOR_MAP        — path to status_color_map.json (default: config/status_color_map.json)
  SHOPIFY_SOURCE_XLSX      — source master merged Excel
  SHOPIFY_CHARGEBACK_ORDERS — comma-sep list of order numbers that are active chargebacks

Usage:
    python build_handoff_excel.py --source master.xlsx --output handoff.xlsx
    python build_handoff_excel.py --source master.xlsx --output handoff.xlsx --flag-chargebacks
"""

import csv
import json
import os
import re
import sys
from pathlib import Path
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8')

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

VAULT_ROOT = Path(os.environ.get("SHOPIFY_VAULT_ROOT", "."))
COLOR_MAP_PATH = Path(os.environ.get(
    "SHOPIFY_COLOR_MAP",
    str(Path(__file__).resolve().parents[3] / "config" / "status_color_map.json")
))
CHARGEBACK_ORDERS_RAW = os.environ.get("SHOPIFY_CHARGEBACK_ORDERS", "")
CHARGEBACK_ORDERS = {o.strip() for o in CHARGEBACK_ORDERS_RAW.split(",") if o.strip()}

FILL_CHARGEBACK = PatternFill("solid", fgColor="FFFFB3B3")
FILL_WHITE      = PatternFill("solid", fgColor="FFFFFFFF")
FILL_GRAY       = PatternFill("solid", fgColor="FFF2F2F2")


def load_color_map(path: Path) -> dict:
    """
    Load merchant status color map from JSON config.
    Maps status text (lowercase) → hex fill color.
    See config/status_color_map.json for format.
    """
    if not path.exists():
        print(f"  [WARN] Color map not found: {path} — using defaults only")
        return {}
    with open(path, encoding='utf-8') as f:
        raw = json.load(f)
    return {k.lower(): PatternFill("solid", fgColor=v) for k, v in raw.items()}


def status_fill(status_text: str, color_map: dict):
    """Return the fill for a given status, or None if not mapped."""
    if not status_text:
        return None
    key = status_text.strip().lower()
    for k, fill in color_map.items():
        if k in key:
            return fill
    return None


def write_legend_sheet(wb, color_map: dict, chargeback_orders: set, total_processed: int):
    """Prepend a legend sheet explaining the color coding."""
    if "Legend" in wb.sheetnames:
        del wb["Legend"]

    legend_ws = wb.create_sheet("Legend", 0)
    legend_ws.column_dimensions["A"].width = 30
    legend_ws.column_dimensions["B"].width = 50

    bold_font = Font(bold=True, size=11)
    header_fill = PatternFill("solid", fgColor="FFDDDDDD")

    # Header row
    ca = legend_ws.cell(1, 1, "Status / Color")
    cb = legend_ws.cell(1, 2, "Meaning")
    ca.font = bold_font; ca.fill = header_fill
    cb.font = bold_font; cb.fill = header_fill

    row = 2
    legend_ws.cell(row, 1, "Light Red (#FFB3B3)")
    legend_ws.cell(row, 2, "CHARGEBACK — Do not process until resolved")
    legend_ws.cell(row, 1).fill = FILL_CHARGEBACK
    row += 1

    for status_key, fill in color_map.items():
        legend_ws.cell(row, 1, status_key.title())
        legend_ws.cell(row, 2, f"Merchant status: {status_key}")
        legend_ws.cell(row, 1).fill = fill
        row += 1

    row += 1
    legend_ws.cell(row, 1, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    row += 1
    legend_ws.cell(row, 1, f"Orders processed: {total_processed}")
    if chargeback_orders:
        row += 1
        legend_ws.cell(row, 1, "Chargeback orders:")
        legend_ws.cell(row, 2, "  ".join(sorted(chargeback_orders)))


def color_sheet(ws, color_map: dict, chargeback_orders: set) -> dict:
    """Apply row coloring to a worksheet. Returns stats."""
    stats = {"chargeback": 0, "status_colored": 0, "plain": 0}
    status_col = None

    # Find a "status" column by header
    for c in range(1, ws.max_column + 1):
        header = str(ws.cell(1, c).value or "").strip().lower()
        if "status" in header:
            status_col = c
            break

    for r in range(2, ws.max_row + 1):
        order_val = str(ws.cell(r, 1).value or "").strip()
        if not order_val:
            continue
        norm = order_val if order_val.startswith("#") else f"#{order_val}"
        is_cb = norm in chargeback_orders

        if is_cb:
            for c in range(1, ws.max_column + 1):
                ws.cell(r, c).fill = FILL_CHARGEBACK
            stats["chargeback"] += 1
        else:
            row_fill = FILL_WHITE if r % 2 == 0 else FILL_GRAY
            for c in range(1, ws.max_column + 1):
                ws.cell(r, c).fill = row_fill

            if status_col:
                status_val = str(ws.cell(r, status_col).value or "").strip()
                sf = status_fill(status_val, color_map)
                if sf:
                    ws.cell(r, status_col).fill = sf
                    stats["status_colored"] += 1
                else:
                    stats["plain"] += 1
            else:
                stats["plain"] += 1

    return stats


def main():
    import argparse
    parser = argparse.ArgumentParser(
        description="Build merchant handoff Excel with production status coloring."
    )
    parser.add_argument("--source", required=True, help="Source master Excel path")
    parser.add_argument("--output", required=True, help="Output Excel path")
    parser.add_argument("--color-map", default=str(COLOR_MAP_PATH), help="Path to status_color_map.json")
    parser.add_argument("--chargebacks", default=CHARGEBACK_ORDERS_RAW,
                        help="Comma-sep chargeback order numbers")
    args = parser.parse_args()

    chargeback_orders = {o.strip() for o in args.chargebacks.split(",") if o.strip()}
    color_map = load_color_map(Path(args.color_map))

    print(f"Loading source Excel: {args.source}")
    wb = load_workbook(args.source)
    total_rows = sum(ws.max_row - 1 for ws in wb.worksheets if ws.max_row > 1)

    print("Applying color coding...")
    for ws in wb.worksheets:
        stats = color_sheet(ws, color_map, chargeback_orders)
        print(f"  Sheet '{ws.title}': chargebacks={stats['chargeback']}, "
              f"status-colored={stats['status_colored']}, plain={stats['plain']}")

    write_legend_sheet(wb, color_map, chargeback_orders, total_rows)

    print(f"\nSaving → {args.output}")
    wb.save(args.output)
    print("Done.")


if __name__ == "__main__":
    main()
