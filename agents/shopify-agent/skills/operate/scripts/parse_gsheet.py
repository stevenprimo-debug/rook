#!/usr/bin/env python3
"""
Generic Shopify operate-mode helper. Sanitized from a prior client engagement.

Parse a merchant's production tracking spreadsheet (Google Sheet export,
xlsx, or csv) into a canonical CSV format for downstream processing.

Supports:
  - Google Sheet markdown table export (via MCP file read → JSON)
  - Direct .xlsx or .csv file read

The column structure is configurable via environment variables to support
any merchant's spreadsheet layout.

Configuration via environment variables (see .env.example):
  SHOPIFY_VAULT_ROOT           — base path for this merchant's data folder
  GSHEET_COL_ORDER             — 0-based column index for order number (default: 0)
  GSHEET_COL_PROD_NOTES        — column index for production notes (default: 3)
  GSHEET_COL_STATUS            — column index for status (default: 4)
  GSHEET_COL_CUSTOM_FIELD      — column index for merchant custom field (default: 5)
  GSHEET_COL_EXTRA             — column index for extra notes (default: 6)
  GSHEET_ORDER_PATTERN         — regex pattern for valid order numbers (default: ^#\d+$)
  GSHEET_INPUT_FORMAT          — json_mcp | xlsx | csv (default: csv)

Usage:
    python parse_gsheet.py --input path/to/input --output path/to/output.csv
    python parse_gsheet.py --input tool_result.json --format json_mcp
"""

import json
import os
import re
import csv
import sys
from pathlib import Path

sys.stdout.reconfigure(encoding='utf-8')

VAULT_ROOT = Path(os.environ.get("SHOPIFY_VAULT_ROOT", "."))

COL_ORDER       = int(os.environ.get("GSHEET_COL_ORDER", 0))
COL_PROD_NOTES  = int(os.environ.get("GSHEET_COL_PROD_NOTES", 3))
COL_STATUS      = int(os.environ.get("GSHEET_COL_STATUS", 4))
COL_CUSTOM      = int(os.environ.get("GSHEET_COL_CUSTOM_FIELD", 5))
COL_EXTRA       = int(os.environ.get("GSHEET_COL_EXTRA", 6))
ORDER_PATTERN   = re.compile(os.environ.get("GSHEET_ORDER_PATTERN", r"^#\d+$"))
INPUT_FORMAT    = os.environ.get("GSHEET_INPUT_FORMAT", "csv")


def parse_markdown_table(content: str) -> list[dict]:
    """
    Parse a markdown pipe-table to rows.
    Returns list of dicts with keys: order, prod_notes, status, custom_field, extra
    """
    rows = []
    for line in content.split('\n'):
        line = line.strip()
        if not line.startswith('|'):
            continue
        cols = [c.strip() for c in line.split('|')]
        cols = [c for c in cols if c and c != ':-:']
        if len(cols) < 3:
            continue
        order_raw = cols[COL_ORDER].replace('\\', '').strip()
        if not ORDER_PATTERN.match(order_raw):
            continue
        rows.append({
            'order':       order_raw,
            'prod_notes':  cols[COL_PROD_NOTES].strip() if len(cols) > COL_PROD_NOTES else '',
            'status':      cols[COL_STATUS].strip() if len(cols) > COL_STATUS else '',
            'custom_field': cols[COL_CUSTOM].strip() if len(cols) > COL_CUSTOM else '',
            'extra':       cols[COL_EXTRA].strip() if len(cols) > COL_EXTRA else '',
        })
    return rows


def parse_json_mcp(json_path: str) -> list[dict]:
    """
    Parse a JSON tool result from MCP file read (fileContent is the markdown table).
    """
    with open(json_path, encoding='utf-8') as f:
        data = json.load(f)
    content = data.get('fileContent', '')
    return parse_markdown_table(content)


def parse_csv_file(csv_path: str) -> list[dict]:
    """
    Parse a direct CSV export with assumed column layout.
    Adjust COL_* env vars to match the merchant's CSV structure.
    """
    rows = []
    with open(csv_path, encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        header = next(reader, None)
        for line in reader:
            if not line:
                continue
            order_raw = line[COL_ORDER].strip() if len(line) > COL_ORDER else ''
            if not order_raw.startswith('#'):
                order_raw = f"#{order_raw}" if order_raw else ''
            if not ORDER_PATTERN.match(order_raw):
                continue
            rows.append({
                'order':       order_raw,
                'prod_notes':  line[COL_PROD_NOTES].strip() if len(line) > COL_PROD_NOTES else '',
                'status':      line[COL_STATUS].strip() if len(line) > COL_STATUS else '',
                'custom_field': line[COL_CUSTOM].strip() if len(line) > COL_CUSTOM else '',
                'extra':       line[COL_EXTRA].strip() if len(line) > COL_EXTRA else '',
            })
    return rows


def parse_xlsx_file(xlsx_path: str) -> list[dict]:
    """Parse an xlsx spreadsheet with assumed column layout."""
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    first = True
    for row in ws.iter_rows(values_only=True):
        if first:
            first = False
            continue  # skip header
        if not row or len(row) <= COL_ORDER:
            continue
        order_raw = str(row[COL_ORDER] or '').strip()
        if not order_raw.startswith('#'):
            order_raw = f"#{order_raw}" if order_raw else ''
        if not ORDER_PATTERN.match(order_raw):
            continue
        rows.append({
            'order':       order_raw,
            'prod_notes':  str(row[COL_PROD_NOTES] or '').strip() if len(row) > COL_PROD_NOTES else '',
            'status':      str(row[COL_STATUS] or '').strip() if len(row) > COL_STATUS else '',
            'custom_field': str(row[COL_CUSTOM] or '').strip() if len(row) > COL_CUSTOM else '',
            'extra':       str(row[COL_EXTRA] or '').strip() if len(row) > COL_EXTRA else '',
        })
    wb.close()
    return rows


def save_canonical_csv(rows: list[dict], out_path: str) -> None:
    fields = ['order', 'prod_notes', 'status', 'custom_field', 'extra']
    with open(out_path, 'w', newline='', encoding='utf-8') as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(rows)
    print(f"Saved canonical CSV: {out_path}")
    if rows:
        print(f"Parsed: {len(rows)} rows  ({rows[0]['order']} to {rows[-1]['order']})")


def main():
    import argparse
    parser = argparse.ArgumentParser(
        description="Parse merchant production spreadsheet to canonical CSV."
    )
    parser.add_argument("--input", required=True, help="Input file path")
    parser.add_argument("--output", required=True, help="Output CSV path")
    parser.add_argument(
        "--format", default=INPUT_FORMAT,
        choices=["json_mcp", "csv", "xlsx"],
        help="Input format (default: from GSHEET_INPUT_FORMAT env)"
    )
    args = parser.parse_args()

    fmt = args.format
    if fmt == "json_mcp":
        rows = parse_json_mcp(args.input)
    elif fmt == "xlsx":
        rows = parse_xlsx_file(args.input)
    else:
        rows = parse_csv_file(args.input)

    save_canonical_csv(rows, args.output)


if __name__ == "__main__":
    main()
