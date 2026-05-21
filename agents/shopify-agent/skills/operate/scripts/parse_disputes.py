#!/usr/bin/env python3
"""
Generic Shopify operate-mode helper. Sanitized from a prior client engagement.

Download bulk operation JSONL, parse disputes, flag chargeback orders.
Cross-references against the master merged Excel.

DUAL-PATH CHARGEBACK DETECTION — see KNOWN_ISSUES.md KI-001:
  Primary:  bulk_disputes.jsonl → disputes table (when Shopify Payments active)
  Fallback: CSV financial_status='chargeback' → orders.has_chargeback flag

Configuration via environment variables (see .env.example):
  SHOPIFY_VAULT_ROOT   — base path for this merchant's data folder
  SHOPIFY_DISPUTES_URL — signed GCS URL for bulk disputes JSONL (expires; see KI-002)

Usage:
    python parse_disputes.py [--url URL] [--local path/to/bulk.jsonl]
"""

import json
import os
import sys
import urllib.request
from datetime import datetime
from pathlib import Path


VAULT_ROOT = Path(os.environ.get("SHOPIFY_VAULT_ROOT", "."))
LOCAL_JSONL_DEFAULT = VAULT_ROOT / "bulk_disputes.jsonl"
DISPUTES_URL = os.environ.get("SHOPIFY_DISPUTES_URL", "")


def download_jsonl(url: str, dest: str) -> None:
    if not url:
        raise ValueError(
            "SHOPIFY_DISPUTES_URL env var not set. Request a new bulk export via API. "
            "See KNOWN_ISSUES.md KI-002 — signed URLs expire."
        )
    print(f"Downloading bulk JSONL...")
    urllib.request.urlretrieve(url, dest)
    print(f"  Saved to: {dest}")


def parse_disputes(jsonl_path: str) -> dict:
    """
    Parse the bulk JSONL. Each line is an order node.
    The disputes field is directly on the order.
    Returns {order_name: [dispute_dicts]}

    NOTE: Returns empty dict when Shopify Payments is paused — see KNOWN_ISSUES.md KI-001.
    Use get_chargeback_orders() from db.py as the fallback path.
    """
    disputes_map = {}
    total = 0
    with open(jsonl_path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            obj = json.loads(line)
            total += 1

            name = obj.get("name", "")
            disputes = obj.get("disputes", [])

            if disputes:
                disputes_map[name] = disputes

    print(f"  Parsed {total} order records")
    print(f"  Orders with disputes field populated: {len(disputes_map)}")
    return disputes_map


def summarize_disputes(disputes_map: dict) -> None:
    """Print a breakdown of dispute statuses."""
    from collections import Counter
    all_statuses = []
    all_types = []

    print("\n=== DISPUTE SUMMARY ===")
    for order_name, disputes in sorted(disputes_map.items()):
        for d in disputes:
            all_statuses.append(d.get("status", "unknown"))
            all_types.append(d.get("initiatedAs", "unknown"))
            print(f"  {order_name}  type={d.get('initiatedAs','?')}  status={d.get('status','?')}")

    print(f"\nStatus counts:")
    for status, count in Counter(all_statuses).most_common():
        print(f"  {status}: {count}")

    print(f"\nInitiatedAs counts:")
    for t, count in Counter(all_types).most_common():
        print(f"  {t}: {count}")


def flag_in_excel(disputes_map: dict, excel_path: str) -> str:
    """
    Read existing master Excel, add a 'Dispute' column, flag matching orders.
    Saves a new file and returns the output path.
    """
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    print(f"\nLoading Excel: {excel_path}")
    wb = load_workbook(excel_path)

    ACTIVE_STATUSES = {"NEEDS_RESPONSE", "UNDER_REVIEW", "ACCEPTED"}
    fill_dispute = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    fill_won     = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")

    for ws in wb.worksheets:
        max_col = ws.max_column
        dispute_col = max_col + 1
        header_cell = ws.cell(row=1, column=dispute_col, value="Dispute")
        header_cell.fill = PatternFill(start_color="1C1C1C", end_color="1C1C1C", fill_type="solid")
        header_cell.font = Font(bold=True, color="FFFFFF")
        header_cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.column_dimensions[get_column_letter(dispute_col)].width = 18

        flagged = 0
        for row_idx in range(2, ws.max_row + 1):
            order_cell = ws.cell(row=row_idx, column=1)
            order_name = str(order_cell.value or "").strip()
            cell = ws.cell(row=row_idx, column=dispute_col)
            cell.alignment = Alignment(vertical="center", wrap_text=False)

            if order_name in disputes_map:
                disputes = disputes_map[order_name]
                statuses = [d.get("status", "") for d in disputes]
                types = [d.get("initiatedAs", "") for d in disputes]
                label_parts = [f"{t}/{s}" for t, s in zip(types, statuses)]
                label = "; ".join(label_parts)
                cell.value = label

                if any(s in ACTIVE_STATUSES for s in statuses):
                    cell.fill = fill_dispute
                    cell.font = Font(color="FFFFFF", bold=True)
                elif any(s == "WON" for s in statuses):
                    cell.fill = fill_won
                    cell.font = Font(color="FFFFFF")
                else:
                    cell.value = label
                flagged += 1

        print(f"  Sheet '{ws.title}': {flagged} orders flagged")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = excel_path.replace(".xlsx", f"_disputes_{timestamp}.xlsx")
    wb.save(out_path)
    print(f"\nSaved: {out_path}")
    return out_path


def find_latest_excel(vault_root: Path, pattern: str = "master_merged_*.xlsx") -> str:
    """Find the most recent Excel file matching pattern in vault_root."""
    import glob
    matches = sorted(glob.glob(str(vault_root / pattern)))
    if not matches:
        raise FileNotFoundError(f"No {pattern} found in {vault_root}")
    latest = matches[-1]
    print(f"  Found Excel: {latest}")
    return latest


def main():
    import argparse
    parser = argparse.ArgumentParser(
        description="Parse Shopify disputes bulk JSONL and flag orders in Excel."
    )
    parser.add_argument("--url", type=str, default="", help="Signed GCS URL for bulk JSONL")
    parser.add_argument("--local", type=str, default="", help="Path to local bulk JSONL file")
    parser.add_argument("--excel", type=str, default="", help="Path to master Excel file")
    args = parser.parse_args()

    local_path = args.local or str(LOCAL_JSONL_DEFAULT)

    # Step 1: Obtain JSONL
    if not os.path.exists(local_path):
        url = args.url or DISPUTES_URL
        download_jsonl(url, local_path)
    else:
        print(f"Using cached JSONL: {local_path}")

    # Step 2: Parse
    disputes_map = parse_disputes(local_path)

    # Step 3: Summarize
    if disputes_map:
        summarize_disputes(disputes_map)
    else:
        print("\nNo disputes found in bulk data.")
        print("  KNOWN ISSUE: Shopify Payments may be paused — disputes API returns empty arrays.")
        print("  See KNOWN_ISSUES.md KI-001 for the dual-path workaround.")
        print("  Use CSV fallback: filter orders where financial_status = 'chargeback'")
        print("  and set orders.has_chargeback = 1 in the database.")

    # Step 4: Flag in Excel if we have disputes
    if disputes_map:
        try:
            excel_path = args.excel or find_latest_excel(VAULT_ROOT)
            flag_in_excel(disputes_map, excel_path)
        except FileNotFoundError as e:
            print(f"\nSkipping Excel update: {e}")
    else:
        print("\nSkipping Excel update — no dispute data to flag.")
        print("Run the CSV fallback via build_handoff_excel.py with --flag-chargebacks.")


if __name__ == "__main__":
    main()
