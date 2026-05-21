#!/usr/bin/env python3
"""
Generic Shopify operate-mode helper. Sanitized from a prior client engagement.

Enrich merchant production spreadsheet orders with custom field values
from Shopify API + CSV fallback, then batch for downstream processing.

Custom field priority (configurable):
  1. Spreadsheet column already has a value → keep it (merchant-confirmed)
  2. Shopify API line-item property → authoritative for recent orders
  3. Shopify CSV shipping address first token → fallback

Skip for downstream batch generation:
  - Status notes indicating done (configurable via SKIP_NOTES / DONE_NOTES env)
  - Missing custom field after all enrichment attempts

Configuration via environment variables (see .env.example):
  SHOPIFY_VAULT_ROOT        — base path for this merchant's data folder
  SHOPIFY_API_XLSX          — path to Shopify API order pull Excel (from shopify_order_pull.py)
  SHOPIFY_CSV_PATH          — path to Shopify CSV export
  BATCH_SKIP_NOTES          — comma-sep notes that mean skip (default: cancelled,chargeback,do not make)
  BATCH_DONE_NOTES          — comma-sep notes that mean done (default: shipped,rust,paint,rba)

Usage:
    python enrich_orders_and_batch.py --gsheet path/to/canonical.csv --output enriched.csv
    python enrich_orders_and_batch.py --gsheet canonical.csv --output enriched.csv --batch
"""

import sys
import re
import csv
import argparse
from pathlib import Path
import os

sys.stdout.reconfigure(encoding='utf-8')

VAULT_ROOT = Path(os.environ.get("SHOPIFY_VAULT_ROOT", "."))
API_XLSX   = os.environ.get("SHOPIFY_API_XLSX", "")
CSV_PATH   = os.environ.get("SHOPIFY_CSV_PATH", "")

SKIP_NOTES_RAW = os.environ.get("BATCH_SKIP_NOTES", "cancelled,chargeback,do not make")
DONE_NOTES_RAW = os.environ.get("BATCH_DONE_NOTES", "shipped,rust,paint,rba")
SKIP_NOTES = {n.strip().lower() for n in SKIP_NOTES_RAW.split(",")}
DONE_NOTES = {n.strip().lower() for n in DONE_NOTES_RAW.split(",")}

_HOUSE_RE = re.compile(r'^[NSEW]?\d+', re.IGNORECASE)


def load_api_values(xlsx_path: str) -> dict:
    """Load {order_num: custom_field_value} from Shopify API export Excel."""
    if not xlsx_path or not Path(xlsx_path).exists():
        print(f"  [WARN] API xlsx not found: {xlsx_path}")
        return {}
    try:
        from openpyxl import load_workbook
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb.active
        api_map = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            order_num = str(row[0] or "").strip()
            # Column index 8 = custom field (House Number / custom property)
            custom_val = str(row[8] or "").strip()
            if order_num and custom_val:
                api_map[order_num] = custom_val
        wb.close()
        print(f"  API custom field values loaded: {len(api_map)} orders")
        return api_map
    except Exception as e:
        print(f"  [WARN] API xlsx load error: {e}")
        return {}


def extract_address_token(addr: str) -> str:
    """Extract first address token (street number) from shipping address."""
    tokens = (addr or "").strip().lstrip("'").split()
    if not tokens:
        return ""
    t = tokens[0]
    return t if _HOUSE_RE.match(t) else ""


def load_csv_values(csv_path: str) -> dict:
    """Load {order_num: shipping_address_first_token} from Shopify CSV export."""
    if not csv_path or not Path(csv_path).exists():
        print(f"  [WARN] CSV not found: {csv_path}")
        return {}
    from collections import defaultdict
    raw = defaultdict(list)
    try:
        with open(csv_path, encoding='utf-8') as f:
            for row in csv.DictReader(f):
                raw[row["Name"]].append(row)
    except Exception as e:
        print(f"  [WARN] CSV load error: {e}")
        return {}
    result = {}
    for name, rows in raw.items():
        addr = rows[0].get("Shipping Address1", "")
        token = extract_address_token(addr)
        if token:
            result[name] = token
    print(f"  CSV shipping address tokens: {len(result)} orders")
    return result


def is_blank(val: str) -> bool:
    v = val.strip().lower()
    return (
        not v
        or v in ("0", "unfulfilled", "n/a", "na", "none", "")
        or v.startswith("waiting")
        or v.startswith("need")
    )


def should_skip(prod_notes: str, custom_val: str) -> tuple[bool, str]:
    notes_lower = (prod_notes or "").lower()
    val_lower   = (custom_val or "").lower()
    for kw in SKIP_NOTES:
        if kw in notes_lower or kw in val_lower:
            return True, f"skip: {kw}"
    return False, ""


def is_done(prod_notes: str) -> bool:
    notes_lower = (prod_notes or "").lower()
    return any(kw in notes_lower for kw in DONE_NOTES)


def enrich(canonical_csv: Path, api_map: dict, csv_values: dict) -> list:
    rows = []
    with open(canonical_csv, encoding='utf-8') as f:
        for r in csv.DictReader(f):
            rows.append(dict(r))

    api_filled  = 0
    csv_filled  = 0
    already_had = 0
    still_blank = 0

    for r in rows:
        order   = r["order"]
        current = r.get("custom_field", "").strip()
        r["custom_field_source"] = "spreadsheet"

        if is_blank(current):
            if order in api_map:
                r["custom_field"] = api_map[order]
                r["custom_field_source"] = "shopify_api"
                api_filled += 1
            elif order in csv_values:
                r["custom_field"] = csv_values[order]
                r["custom_field_source"] = "csv_addr"
                csv_filled += 1
            else:
                r["custom_field_source"] = "missing"
                still_blank += 1
        else:
            already_had += 1

    print(f"\nEnrichment result:")
    print(f"  Already had custom field: {already_had}")
    print(f"  Filled from Shopify API: {api_filled}")
    print(f"  Filled from CSV address: {csv_filled}")
    print(f"  Still missing:           {still_blank}")
    return rows


def save_enriched(rows: list, out_path: str) -> None:
    fields = ["order", "prod_notes", "status", "custom_field", "extra", "custom_field_source"]
    with open(out_path, 'w', newline='', encoding='utf-8') as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction='ignore')
        w.writeheader()
        w.writerows(rows)
    print(f"\nSaved enriched CSV: {out_path}")


def build_batch_list(rows: list) -> list:
    """Filter enriched rows to the set ready for downstream batch processing."""
    batch = []
    skipped = 0
    done = 0
    missing = 0

    for r in rows:
        val = r.get("custom_field", "").strip()
        if not val or is_blank(val):
            missing += 1
            continue
        skip, reason = should_skip(r.get("prod_notes", ""), val)
        if skip:
            skipped += 1
            continue
        if is_done(r.get("prod_notes", "")):
            done += 1
            continue
        batch.append({
            "order_id":  r["order"].lstrip("#"),
            "order_num": r["order"],
            "custom_field": val,
            "custom_field_source": r.get("custom_field_source", ""),
            "prod_notes": r.get("prod_notes", ""),
        })

    print(f"\nBatch list: {len(batch)} orders ready")
    print(f"  Skipped (cancel/chargeback): {skipped}")
    print(f"  Already done:               {done}")
    print(f"  Missing custom field:        {missing}")
    return batch


def main():
    parser = argparse.ArgumentParser(
        description="Enrich merchant spreadsheet orders with Shopify API + CSV fallback."
    )
    parser.add_argument("--gsheet", required=True, help="Path to canonical CSV (from parse_gsheet.py)")
    parser.add_argument("--output", required=True, help="Path to save enriched CSV")
    parser.add_argument("--api-xlsx", default=API_XLSX, help="Shopify API order pull Excel")
    parser.add_argument("--csv", default=CSV_PATH, help="Shopify CSV export path")
    parser.add_argument("--batch", action="store_true", help="Print batch-ready order list")
    args = parser.parse_args()

    print("Loading Shopify API values...")
    api_map = load_api_values(args.api_xlsx)

    print("Loading CSV shipping address tokens...")
    csv_values = load_csv_values(args.csv)

    print(f"\nEnriching {Path(args.gsheet).name}...")
    rows = enrich(Path(args.gsheet), api_map, csv_values)
    save_enriched(rows, args.output)

    if args.batch:
        batch = build_batch_list(rows)
        print(f"\nBatch-ready orders:")
        for o in batch[:10]:
            print(f"  {o['order_num']}  custom_field={o['custom_field']!r}  src={o['custom_field_source']}")
        if len(batch) > 10:
            print(f"  ... and {len(batch) - 10} more")


if __name__ == "__main__":
    main()
